"""
Unit tests for reporting.excel_generator (ExcelGenerator)

Tests report generation logic with real Resource/DiscoveryResult objects.
"""
import os
import pytest
import tempfile
from unittest.mock import patch
from datetime import datetime

from reporting.excel_generator import ExcelGenerator
from resource_discovery.models import Resource, DiscoveryResult, DiscoverySource
from tests.conftest import make_resource, make_discovery_result


# ===================================================================
# Helpers
# ===================================================================

@pytest.fixture
def tmp_output_dir(tmp_path):
    """Use pytest's tmp_path for output directory."""
    return str(tmp_path / "reports")


@pytest.fixture
def generator(tmp_output_dir):
    return ExcelGenerator(output_dir=tmp_output_dir)


@pytest.fixture
def populated_result(sample_resources):
    return make_discovery_result(
        resources=sample_resources,
        total_count=len(sample_resources),
    )


# ===================================================================
# generate_report
# ===================================================================

class TestGenerateReport:

    def test_creates_file(self, generator, populated_result):
        path = generator.generate_report(populated_result)
        assert path != ""
        assert os.path.exists(path)
        assert path.endswith(".xlsx")

    def test_empty_result_returns_empty_string(self, generator):
        result = make_discovery_result(resources=[], total_count=0)
        path = generator.generate_report(result)
        assert path == ""

    def test_default_filename_has_timestamp(self, generator, populated_result):
        path = generator.generate_report(populated_result)
        basename = os.path.basename(path)
        assert basename.startswith("CloudAuditor_Report_")
        assert basename.endswith(".xlsx")

    def test_custom_filename(self, generator, populated_result):
        path = generator.generate_report(populated_result, filename="custom_report.xlsx")
        assert os.path.basename(path) == "custom_report.xlsx"

    def test_output_dir_created(self, tmp_path):
        new_dir = str(tmp_path / "nonexistent" / "reports")
        gen = ExcelGenerator(output_dir=new_dir)
        assert os.path.isdir(new_dir)

    def test_file_has_content(self, generator, populated_result):
        path = generator.generate_report(populated_result)
        assert os.path.getsize(path) > 0


# ===================================================================
# _create_summary_stats
# ===================================================================

class TestCreateSummaryStats:

    def test_summary_has_header_rows(self, generator, sample_resources):
        import pandas as pd
        data = [r.to_dict() for r in sample_resources]
        df = pd.DataFrame(data)

        summary = generator._create_summary_stats(df)

        # First two rows should be the header (Total Accounts, Total Resources)
        assert summary.iloc[0]["resource_type"] == "Total Accounts"
        assert summary.iloc[1]["resource_type"] == "Total Resources"

    def test_summary_total_resources_correct(self, generator, sample_resources):
        import pandas as pd
        data = [r.to_dict() for r in sample_resources]
        df = pd.DataFrame(data)

        summary = generator._create_summary_stats(df)

        total_row = summary[summary["resource_type"] == "Total Resources"]
        assert total_row.iloc[0]["count"] == len(sample_resources)

    def test_summary_total_accounts_correct(self, generator, sample_resources):
        import pandas as pd
        data = [r.to_dict() for r in sample_resources]
        df = pd.DataFrame(data)

        summary = generator._create_summary_stats(df)

        accounts_row = summary[summary["resource_type"] == "Total Accounts"]
        # sample_resources has 2 unique accounts (123456789012, 987654321098)
        assert accounts_row.iloc[0]["count"] == 2


# ===================================================================
# Sheet structure (verify Excel internals)
# ===================================================================

class TestSheetStructure:

    def test_has_executive_summary_sheet(self, generator, populated_result):
        import openpyxl
        path = generator.generate_report(populated_result)
        wb = openpyxl.load_workbook(path)
        assert "Executive Summary" in wb.sheetnames

    def test_has_service_tabs(self, generator, populated_result):
        """Each unique service prefix should get its own sheet."""
        import openpyxl
        path = generator.generate_report(populated_result)
        wb = openpyxl.load_workbook(path)

        # sample_resources has resource_types with :: delimiters
        # The generator splits on ':' → first part
        # AWS::EC2::Instance → "AWS" (it's the first part before any ':')
        sheet_names = wb.sheetnames
        # At minimum, should have Executive Summary + at least 1 service tab
        assert len(sheet_names) >= 2

    def test_sheet_name_truncation(self, generator):
        """Sheet names over 31 chars should be truncated."""
        long_type_resource = make_resource(
            resource_type="very_long_service_name_that_exceeds_limit:subtype",
        )
        result = make_discovery_result(
            resources=[long_type_resource],
            total_count=1,
        )
        path = generator.generate_report(result)

        import openpyxl
        wb = openpyxl.load_workbook(path)
        for name in wb.sheetnames:
            assert len(name) <= 31
